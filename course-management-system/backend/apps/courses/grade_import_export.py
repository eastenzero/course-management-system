"""
成绩导入导出服务
"""

from django.http import HttpResponse
from django.core.exceptions import ValidationError
from django.db import transaction
from typing import Dict, List, Any, Tuple
import csv
from io import StringIO, BytesIO
from decimal import Decimal

from .models import Course, Enrollment, Grade, GradeComponent
from apps.users.models import User


class GradeImportExportService:
    """成绩导入导出服务"""
    
    @staticmethod
    def export_grades_to_excel(course_id: int, options: Dict[str, Any] = None) -> bytes:
        """导出成绩为Excel格式"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("需要安装 openpyxl 库来支持Excel导出")
        
        options = options or {}
        include_details = options.get('include_details', True)
        include_statistics = options.get('include_statistics', True)
        
        try:
            course = Course.objects.get(id=course_id)
        except Course.DoesNotExist:
            raise ValueError('课程不存在')
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{course.name}-成绩单"
        
        # 设置样式
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 课程信息
        ws['A1'] = f"课程名称: {course.name}"
        ws['A2'] = f"课程代码: {course.code}"
        ws['A3'] = f"学期: {course.semester}"
        ws['A4'] = f"学分: {course.credits}"
        
        # 获取选课记录
        enrollments = course.enrollments.filter(is_active=True).select_related(
            'student', 'student__student_profile'
        ).order_by('student__username')
        
        # 获取成绩组成
        components = course.grade_components.all().order_by('order')
        
        # 设置表头
        start_row = 6
        headers = ['学号', '姓名', '班级']
        
        if include_details and components.exists():
            # 包含详细成绩
            for component in components:
                headers.append(f"{component.name}({component.weight}%)")
        
        headers.extend(['总成绩', '等级'])
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # 写入数据
        row = start_row + 1
        for enrollment in enrollments:
            student = enrollment.student
            
            # 基本信息
            ws.cell(row=row, column=1, value=student.username)
            ws.cell(row=row, column=2, value=f"{student.first_name} {student.last_name}".strip() or student.username)
            
            # 班级信息
            class_name = ''
            if hasattr(student, 'student_profile') and student.student_profile:
                class_name = student.student_profile.class_name
            ws.cell(row=row, column=3, value=class_name)
            
            col = 4
            
            # 详细成绩
            if include_details and components.exists():
                for component in components:
                    component_grades = enrollment.detailed_grades.filter(component=component)
                    if component_grades.exists():
                        avg_score = sum(grade.percentage_score for grade in component_grades) / component_grades.count()
                        ws.cell(row=row, column=col, value=round(avg_score, 2))
                    else:
                        ws.cell(row=row, column=col, value='--')
                    col += 1
            
            # 总成绩和等级
            ws.cell(row=row, column=col, value=float(enrollment.score) if enrollment.score else '--')
            ws.cell(row=row, column=col + 1, value=enrollment.grade or '--')
            
            # 设置边框
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).border = border
            
            row += 1
        
        # 统计信息
        if include_statistics and enrollments.exists():
            stats_row = row + 2
            ws.cell(row=stats_row, column=1, value="统计信息").font = Font(bold=True)
            
            # 计算统计数据
            scores = [float(e.score) for e in enrollments if e.score is not None]
            if scores:
                ws.cell(row=stats_row + 1, column=1, value=f"总人数: {enrollments.count()}")
                ws.cell(row=stats_row + 2, column=1, value=f"平均分: {round(sum(scores) / len(scores), 2)}")
                ws.cell(row=stats_row + 3, column=1, value=f"最高分: {max(scores)}")
                ws.cell(row=stats_row + 4, column=1, value=f"最低分: {min(scores)}")
                ws.cell(row=stats_row + 5, column=1, value=f"及格率: {round(len([s for s in scores if s >= 60]) / len(scores) * 100, 2)}%")
        
        # 调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    @staticmethod
    def export_grades_to_csv(course_id: int, options: Dict[str, Any] = None) -> str:
        """导出成绩为CSV格式"""
        options = options or {}
        include_details = options.get('include_details', True)
        
        try:
            course = Course.objects.get(id=course_id)
        except Course.DoesNotExist:
            raise ValueError('课程不存在')
        
        output = StringIO()
        writer = csv.writer(output)
        
        # 课程信息
        writer.writerow([f"课程名称: {course.name}"])
        writer.writerow([f"课程代码: {course.code}"])
        writer.writerow([f"学期: {course.semester}"])
        writer.writerow([])  # 空行
        
        # 获取数据
        enrollments = course.enrollments.filter(is_active=True).select_related(
            'student', 'student__student_profile'
        ).order_by('student__username')
        
        components = course.grade_components.all().order_by('order')
        
        # 表头
        headers = ['学号', '姓名', '班级']
        
        if include_details and components.exists():
            for component in components:
                headers.append(f"{component.name}({component.weight}%)")
        
        headers.extend(['总成绩', '等级'])
        writer.writerow(headers)
        
        # 数据行
        for enrollment in enrollments:
            student = enrollment.student
            row = [
                student.username,
                f"{student.first_name} {student.last_name}".strip() or student.username,
                getattr(student.student_profile, 'class_name', '') if hasattr(student, 'student_profile') and student.student_profile else ''
            ]
            
            # 详细成绩
            if include_details and components.exists():
                for component in components:
                    component_grades = enrollment.detailed_grades.filter(component=component)
                    if component_grades.exists():
                        avg_score = sum(grade.percentage_score for grade in component_grades) / component_grades.count()
                        row.append(round(avg_score, 2))
                    else:
                        row.append('--')
            
            # 总成绩和等级
            row.append(float(enrollment.score) if enrollment.score else '--')
            row.append(enrollment.grade or '--')
            
            writer.writerow(row)
        
        return output.getvalue()
    
    @staticmethod
    def validate_import_data(data: List[Dict[str, Any]], course_id: int) -> Tuple[bool, List[str]]:
        """验证导入数据"""
        errors = []
        
        try:
            course = Course.objects.get(id=course_id)
        except Course.DoesNotExist:
            errors.append('课程不存在')
            return False, errors
        
        if not data:
            errors.append('导入数据为空')
            return False, errors
        
        # 检查必需字段
        required_fields = ['student_id', 'score']
        
        for i, row in enumerate(data):
            row_num = i + 1
            
            # 检查必需字段
            for field in required_fields:
                if field not in row or not row[field]:
                    errors.append(f'第{row_num}行缺少必需字段: {field}')
            
            # 验证学号
            if 'student_id' in row:
                try:
                    student = User.objects.get(username=row['student_id'], user_type='student')
                    # 检查是否选课
                    if not course.enrollments.filter(student=student, is_active=True).exists():
                        errors.append(f'第{row_num}行学生{row["student_id"]}未选择此课程')
                except User.DoesNotExist:
                    errors.append(f'第{row_num}行学生{row["student_id"]}不存在')
            
            # 验证成绩
            if 'score' in row:
                try:
                    score = float(row['score'])
                    if score < 0 or score > 100:
                        errors.append(f'第{row_num}行成绩{score}超出有效范围(0-100)')
                except (ValueError, TypeError):
                    errors.append(f'第{row_num}行成绩格式错误: {row["score"]}')
        
        return len(errors) == 0, errors
    
    @staticmethod
    def import_grades_from_data(data: List[Dict[str, Any]], course_id: int) -> Tuple[int, List[str]]:
        """从导入数据创建/更新成绩"""
        try:
            course = Course.objects.get(id=course_id)
        except Course.DoesNotExist:
            return 0, ['课程不存在']
        
        updated_count = 0
        errors = []
        
        with transaction.atomic():
            for i, row in enumerate(data):
                try:
                    student_id = row.get('student_id')
                    score = float(row.get('score'))
                    
                    # 获取学生和选课记录
                    student = User.objects.get(username=student_id, user_type='student')
                    enrollment = course.enrollments.get(student=student, is_active=True)
                    
                    # 更新成绩
                    enrollment.score = Decimal(str(score))
                    
                    # 计算等级
                    if score >= 90:
                        enrollment.grade = 'A'
                    elif score >= 80:
                        enrollment.grade = 'B'
                    elif score >= 70:
                        enrollment.grade = 'C'
                    elif score >= 60:
                        enrollment.grade = 'D'
                    else:
                        enrollment.grade = 'F'
                    
                    enrollment.save()
                    updated_count += 1
                    
                except User.DoesNotExist:
                    errors.append(f'第{i+1}行: 学生{student_id}不存在')
                except Enrollment.DoesNotExist:
                    errors.append(f'第{i+1}行: 学生{student_id}未选择此课程')
                except Exception as e:
                    errors.append(f'第{i+1}行: 更新失败 - {str(e)}')
        
        return updated_count, errors
    
    @staticmethod
    def generate_grade_template(course_id: int) -> bytes:
        """生成成绩导入模板"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ImportError("需要安装 openpyxl 库来支持Excel模板生成")
        
        try:
            course = Course.objects.get(id=course_id)
        except Course.DoesNotExist:
            raise ValueError('课程不存在')
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "成绩导入模板"
        
        # 设置表头
        headers = ['学号', '姓名', '班级', '总成绩', '备注']
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # 填入学生信息
        enrollments = course.enrollments.filter(is_active=True).select_related(
            'student', 'student__student_profile'
        ).order_by('student__username')
        
        for row, enrollment in enumerate(enrollments, 2):
            student = enrollment.student
            ws.cell(row=row, column=1, value=student.username)
            ws.cell(row=row, column=2, value=f"{student.first_name} {student.last_name}".strip() or student.username)
            
            class_name = ''
            if hasattr(student, 'student_profile') and student.student_profile:
                class_name = student.student_profile.class_name
            ws.cell(row=row, column=3, value=class_name)
            
            # 总成绩列留空供填写
            ws.cell(row=row, column=4, value='')
            ws.cell(row=row, column=5, value='')
        
        # 调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # 添加说明
        ws.cell(row=len(enrollments) + 4, column=1, value="说明:")
        ws.cell(row=len(enrollments) + 5, column=1, value="1. 请在'总成绩'列填入0-100的数值")
        ws.cell(row=len(enrollments) + 6, column=1, value="2. 不要修改学号、姓名、班级列")
        ws.cell(row=len(enrollments) + 7, column=1, value="3. 备注列可选填")
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
