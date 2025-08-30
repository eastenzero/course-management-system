# file: backend/apps/algorithms/views.py
# 功能: 算法API视图

import logging
import json
import os
import sys
from typing import List, Dict, Any

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import JsonResponse, HttpResponse
from django.conf import settings

# 添加算法模块路径
algorithms_path = os.path.join(settings.BASE_DIR, '..', 'algorithms')
sys.path.insert(0, algorithms_path)

try:
    from algorithms.engine import SchedulingEngine, AlgorithmType
    from algorithms.models import Assignment, TeacherPreference
    from algorithms.constraints.manager import ConstraintManager
    from algorithms.conflict.analyzer import ConflictAnalyzer
except ImportError as e:
    logging.error(f"Failed to import algorithms module: {e}")
    SchedulingEngine = None
    AlgorithmType = None

from apps.courses.models import Course
from apps.classrooms.models import Classroom
from apps.users.models import User
from apps.schedules.models import Schedule, TimeSlot
from .serializers import (
    ScheduleGenerationRequestSerializer,
    ScheduleResultSerializer,
    ScheduleAnalysisSerializer,
    AlgorithmStatisticsSerializer,
    ScheduleOptimizationRequestSerializer,
    ConflictResolutionRequestSerializer,
    ScheduleExportRequestSerializer,
    AssignmentSerializer,
    ConflictSerializer,
)

logger = logging.getLogger(__name__)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generate_schedule(request):
    """生成排课方案"""
    if not SchedulingEngine:
        return Response(
            {'error': '算法模块未正确加载'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    
    serializer = ScheduleGenerationRequestSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    data = serializer.validated_data
    
    try:
        # 获取数据
        courses_data = _get_courses_data(
            data.get('course_ids'),
            data['semester'],
            data['academic_year']
        )
        teachers_data = _get_teachers_data(data.get('teacher_ids'))
        classrooms_data = _get_classrooms_data(data.get('classroom_ids'))
        teacher_preferences = _get_teacher_preferences()
        
        # 初始化排课引擎
        engine = SchedulingEngine(
            enable_conflict_resolution=data['enable_conflict_resolution'],
            enable_optimization=data['enable_optimization']
        )
        
        engine.initialize(
            courses_data,
            teachers_data,
            classrooms_data,
            teacher_preferences
        )
        
        # 选择算法
        algorithm_map = {
            'greedy': AlgorithmType.GREEDY,
            'genetic': AlgorithmType.GENETIC,
            'parallel_genetic': AlgorithmType.PARALLEL_GENETIC,
            'hybrid': AlgorithmType.HYBRID,
            'optimizer': AlgorithmType.OPTIMIZER,
        }
        
        algorithm = algorithm_map.get(data['algorithm'], AlgorithmType.HYBRID)
        
        # 生成排课方案
        result = engine.generate_schedule(
            algorithm=algorithm,
            algorithm_params=data.get('algorithm_params', {}),
            validate_result=True
        )
        
        # 转换结果格式
        response_data = _convert_schedule_result(result, courses_data, teachers_data, classrooms_data)
        
        # 保存到数据库（可选）
        if data.get('save_to_database', True):
            _save_schedule_to_database(result, data['semester'], data['academic_year'])
        
        return Response(response_data, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Schedule generation failed: {e}")
        return Response(
            {'error': f'排课生成失败: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def analyze_schedule(request):
    """分析排课方案"""
    if not SchedulingEngine:
        return Response(
            {'error': '算法模块未正确加载'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    
    schedule_id = request.data.get('schedule_id')
    if not schedule_id:
        return Response(
            {'error': '缺少schedule_id参数'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        # 获取排课数据
        schedule_data = _get_schedule_data(schedule_id)
        if not schedule_data:
            return Response(
                {'error': '排课方案不存在'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # 初始化引擎和分析器
        engine = SchedulingEngine()
        courses_data = _get_courses_data()
        teachers_data = _get_teachers_data()
        classrooms_data = _get_classrooms_data()
        
        engine.initialize(courses_data, teachers_data, classrooms_data)
        
        # 分析排课方案
        analysis = engine.analyze_schedule(schedule_data)
        
        return Response(analysis, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Schedule analysis failed: {e}")
        return Response(
            {'error': f'排课分析失败: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def optimize_schedule(request):
    """优化排课方案"""
    if not SchedulingEngine:
        return Response(
            {'error': '算法模块未正确加载'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    
    serializer = ScheduleOptimizationRequestSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    data = serializer.validated_data
    
    try:
        # 获取原始排课数据
        original_schedule = _get_schedule_data(data['schedule_id'])
        if not original_schedule:
            return Response(
                {'error': '排课方案不存在'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # 初始化优化器
        from algorithms.optimizer.schedule_optimizer import ScheduleOptimizer
        from algorithms.constraints.manager import ConstraintManager
        
        constraint_manager = ConstraintManager()
        optimizer = ScheduleOptimizer(
            constraint_manager=constraint_manager,
            max_iterations=data.get('max_iterations', 1000),
            improvement_threshold=data.get('improvement_threshold', 0.01)
        )
        
        # 获取数据
        courses_data = _get_courses_data()
        teachers_data = _get_teachers_data()
        classrooms_data = _get_classrooms_data()
        
        constraint_manager.set_data_cache(courses_data, teachers_data, classrooms_data)
        
        # 执行优化
        optimized_result = optimizer.optimize_schedule(
            original_schedule.assignments,
            teachers_data,
            classrooms_data,
            courses_data
        )
        
        # 转换结果格式
        response_data = _convert_schedule_result(optimized_result, courses_data, teachers_data, classrooms_data)
        
        return Response(response_data, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Schedule optimization failed: {e}")
        return Response(
            {'error': f'排课优化失败: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_algorithm_statistics(request):
    """获取算法统计信息"""
    try:
        # 这里可以从缓存或数据库中获取统计信息
        # 暂时返回模拟数据
        stats = {
            'engine_stats': {
                'total_schedules_generated': 0,
                'successful_schedules': 0,
                'failed_schedules': 0,
                'average_generation_time': 0.0,
                'algorithm_usage': {
                    'greedy': 0,
                    'genetic': 0,
                    'parallel_genetic': 0,
                    'hybrid': 0,
                    'optimizer': 0,
                },
            },
            'constraint_manager_stats': {
                'total_checks': 0,
                'hard_violations': 0,
                'average_soft_score': 0.0,
            },
            'conflict_detector_stats': {
                'total_checks': 0,
                'conflicts_found': 0,
                'teacher_conflicts': 0,
                'classroom_conflicts': 0,
            },
            'conflict_resolver_stats': {
                'conflicts_resolved': 0,
                'conflicts_failed': 0,
                'overall_success_rate': 0.0,
            },
        }
        
        return Response(stats, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Failed to get algorithm statistics: {e}")
        return Response(
            {'error': f'获取统计信息失败: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_schedule(request):
    """导出排课方案"""
    serializer = ScheduleExportRequestSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    data = serializer.validated_data
    
    try:
        # 获取排课数据
        schedule_data = _get_schedule_data(data['schedule_id'])
        if not schedule_data:
            return Response(
                {'error': '排课方案不存在'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        export_format = data['format']
        
        if export_format == 'json':
            response_data = schedule_data.to_dict()
            return Response(response_data, status=status.HTTP_200_OK)
        
        elif export_format == 'excel':
            # 生成Excel文件
            excel_file = _generate_excel_export(schedule_data, data)
            response = HttpResponse(
                excel_file,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="schedule_{data["schedule_id"]}.xlsx"'
            return response
        
        elif export_format == 'csv':
            # 生成CSV文件
            csv_file = _generate_csv_export(schedule_data, data)
            response = HttpResponse(csv_file, content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="schedule_{data["schedule_id"]}.csv"'
            return response
        
        elif export_format == 'pdf':
            # 生成PDF文件
            pdf_file = _generate_pdf_export(schedule_data, data)
            response = HttpResponse(pdf_file, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="schedule_{data["schedule_id"]}.pdf"'
            return response
        
        else:
            return Response(
                {'error': f'不支持的导出格式: {export_format}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
    except Exception as e:
        logger.error(f"Schedule export failed: {e}")
        return Response(
            {'error': f'排课导出失败: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# 辅助函数

def _get_courses_data(course_ids=None, semester=None, academic_year=None):
    """获取课程数据"""
    queryset = Course.objects.filter(is_active=True, is_published=True)

    if course_ids:
        queryset = queryset.filter(id__in=course_ids)

    if semester:
        queryset = queryset.filter(semester=semester)

    if academic_year:
        queryset = queryset.filter(academic_year=academic_year)

    courses_data = []
    for course in queryset:
        # 获取合格教师列表
        qualified_teachers = list(course.teachers.values_list('id', flat=True))

        course_data = {
            'id': course.id,
            'name': course.name,
            'code': course.code,
            'credits': course.credits,
            'max_students': course.max_students,
            'course_type': course.course_type,
            'semester': course.semester,
            'academic_year': course.academic_year,
            'is_active': course.is_active,
            'is_published': course.is_published,
            'qualified_teachers': qualified_teachers,
        }
        courses_data.append(course_data)

    return courses_data


def _get_teachers_data(teacher_ids=None):
    """获取教师数据"""
    queryset = User.objects.filter(user_type='teacher', is_active=True)

    if teacher_ids:
        queryset = queryset.filter(id__in=teacher_ids)

    teachers_data = []
    for teacher in queryset:
        # 获取教师可教授的课程
        qualified_courses = list(teacher.courses.values_list('id', flat=True))

        teacher_data = {
            'id': teacher.id,
            'name': teacher.get_full_name(),
            'username': teacher.username,
            'email': teacher.email,
            'department': getattr(teacher, 'department', ''),
            'max_weekly_hours': getattr(teacher, 'max_weekly_hours', 20),
            'max_daily_hours': getattr(teacher, 'max_daily_hours', 8),
            'qualified_courses': qualified_courses,
        }
        teachers_data.append(teacher_data)

    return teachers_data


def _get_classrooms_data(classroom_ids=None):
    """获取教室数据"""
    queryset = Classroom.objects.filter(is_available=True, is_active=True)

    if classroom_ids:
        queryset = queryset.filter(id__in=classroom_ids)

    classrooms_data = []
    for classroom in queryset:
        classroom_data = {
            'id': classroom.id,
            'name': classroom.name,
            'building': classroom.building,
            'floor': classroom.floor,
            'capacity': classroom.capacity,
            'room_type': classroom.room_type,
            'equipment': classroom.equipment,
            'is_available': classroom.is_available,
            'is_active': classroom.is_active,
        }
        classrooms_data.append(classroom_data)

    return classrooms_data


def _get_teacher_preferences():
    """获取教师时间偏好"""
    # 这里可以从数据库中获取教师偏好数据
    # 暂时返回空列表
    return []


def _convert_schedule_result(result, courses_data, teachers_data, classrooms_data):
    """转换排课结果格式"""
    # 创建查找字典
    course_dict = {c['id']: c for c in courses_data}
    teacher_dict = {t['id']: t for t in teachers_data}
    classroom_dict = {c['id']: c for c in classrooms_data}

    # 星期名称映射
    day_names = ['', '周一', '周二', '周三', '周四', '周五', '周六', '周日']

    # 转换分配
    assignments_data = []
    for assignment in result.assignments:
        course = course_dict.get(assignment.course_id, {})
        teacher = teacher_dict.get(assignment.teacher_id, {})
        classroom = classroom_dict.get(assignment.classroom_id, {})

        assignment_data = {
            'course_id': assignment.course_id,
            'course_name': course.get('name', f'Course {assignment.course_id}'),
            'teacher_id': assignment.teacher_id,
            'teacher_name': teacher.get('name', f'Teacher {assignment.teacher_id}'),
            'classroom_id': assignment.classroom_id,
            'classroom_name': classroom.get('name', f'Classroom {assignment.classroom_id}'),
            'day_of_week': assignment.day_of_week,
            'day_name': day_names[assignment.day_of_week] if assignment.day_of_week < len(day_names) else f'Day {assignment.day_of_week}',
            'time_slot': assignment.time_slot,
            'time_name': f'第{assignment.time_slot}节',
            'semester': assignment.semester,
            'academic_year': assignment.academic_year,
            'week_range': assignment.week_range,
        }
        assignments_data.append(assignment_data)

    # 转换冲突
    conflicts_data = []
    for conflict in result.conflicts:
        conflict_assignments = []
        for assignment in conflict.assignments:
            course = course_dict.get(assignment.course_id, {})
            teacher = teacher_dict.get(assignment.teacher_id, {})
            classroom = classroom_dict.get(assignment.classroom_id, {})

            conflict_assignment = {
                'course_id': assignment.course_id,
                'course_name': course.get('name', f'Course {assignment.course_id}'),
                'teacher_id': assignment.teacher_id,
                'teacher_name': teacher.get('name', f'Teacher {assignment.teacher_id}'),
                'classroom_id': assignment.classroom_id,
                'classroom_name': classroom.get('name', f'Classroom {assignment.classroom_id}'),
                'day_of_week': assignment.day_of_week,
                'day_name': day_names[assignment.day_of_week] if assignment.day_of_week < len(day_names) else f'Day {assignment.day_of_week}',
                'time_slot': assignment.time_slot,
                'time_name': f'第{assignment.time_slot}节',
                'semester': assignment.semester,
                'academic_year': assignment.academic_year,
            }
            conflict_assignments.append(conflict_assignment)

        conflict_data = {
            'conflict_type': conflict.conflict_type,
            'assignments': conflict_assignments,
            'description': conflict.description,
            'severity': conflict.severity,
            'created_at': conflict.created_at.isoformat(),
        }
        conflicts_data.append(conflict_data)

    return {
        'assignments': assignments_data,
        'conflicts': conflicts_data,
        'fitness_score': result.fitness_score,
        'algorithm_used': result.algorithm_used,
        'generation_time': result.generation_time,
        'is_valid': result.is_valid,
        'total_assignments': result.total_assignments,
        'conflict_count': result.conflict_count,
        'created_at': result.created_at.isoformat(),
        'metadata': result.metadata,
    }


def _save_schedule_to_database(result, semester, academic_year):
    """保存排课结果到数据库"""
    try:
        # 创建排课记录
        schedule = Schedule.objects.create(
            name=f"{academic_year}_{semester}_智能排课",
            semester=semester,
            academic_year=academic_year,
            algorithm_used=result.algorithm_used,
            fitness_score=result.fitness_score,
            generation_time=result.generation_time,
            is_published=False,
            metadata=result.metadata,
        )

        # 保存分配记录
        from apps.schedules.models import ScheduleAssignment

        for assignment in result.assignments:
            ScheduleAssignment.objects.create(
                schedule=schedule,
                course_id=assignment.course_id,
                teacher_id=assignment.teacher_id,
                classroom_id=assignment.classroom_id,
                day_of_week=assignment.day_of_week,
                time_slot=assignment.time_slot,
                week_range=assignment.week_range,
            )

        return schedule.id

    except Exception as e:
        logger.error(f"Failed to save schedule to database: {e}")
        return None


def _get_schedule_data(schedule_id):
    """从数据库获取排课数据"""
    try:
        schedule = Schedule.objects.get(id=schedule_id)
        assignments = schedule.assignments.all()

        # 转换为算法模型格式
        algorithm_assignments = []
        for assignment in assignments:
            algorithm_assignment = Assignment(
                course_id=assignment.course_id,
                teacher_id=assignment.teacher_id,
                classroom_id=assignment.classroom_id,
                day_of_week=assignment.day_of_week,
                time_slot=assignment.time_slot,
                semester=schedule.semester,
                academic_year=schedule.academic_year,
                week_range=assignment.week_range,
            )
            algorithm_assignments.append(algorithm_assignment)

        # 创建ScheduleResult对象
        from algorithms.models import ScheduleResult

        result = ScheduleResult(
            assignments=algorithm_assignments,
            conflicts=[],  # 需要重新检测
            fitness_score=schedule.fitness_score,
            algorithm_used=schedule.algorithm_used,
            generation_time=schedule.generation_time,
            metadata=schedule.metadata or {},
        )

        return result

    except Schedule.DoesNotExist:
        return None
    except Exception as e:
        logger.error(f"Failed to get schedule data: {e}")
        return None


def _generate_excel_export(schedule_data, options):
    """生成Excel导出文件"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO

        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "排课表"

        # 设置标题
        ws['A1'] = "课程排课表"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:H1')

        # 设置表头
        headers = ['课程名称', '教师', '教室', '星期', '时间段', '学期', '学年', '周次']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

        # 填充数据
        courses_data = _get_courses_data()
        teachers_data = _get_teachers_data()
        classrooms_data = _get_classrooms_data()

        course_dict = {c['id']: c for c in courses_data}
        teacher_dict = {t['id']: t for t in teachers_data}
        classroom_dict = {c['id']: c for c in classrooms_data}

        day_names = ['', '周一', '周二', '周三', '周四', '周五', '周六', '周日']

        for row, assignment in enumerate(schedule_data.assignments, 4):
            course = course_dict.get(assignment.course_id, {})
            teacher = teacher_dict.get(assignment.teacher_id, {})
            classroom = classroom_dict.get(assignment.classroom_id, {})

            ws.cell(row=row, column=1, value=course.get('name', f'Course {assignment.course_id}'))
            ws.cell(row=row, column=2, value=teacher.get('name', f'Teacher {assignment.teacher_id}'))
            ws.cell(row=row, column=3, value=classroom.get('name', f'Classroom {assignment.classroom_id}'))
            ws.cell(row=row, column=4, value=day_names[assignment.day_of_week] if assignment.day_of_week < len(day_names) else f'Day {assignment.day_of_week}')
            ws.cell(row=row, column=5, value=f'第{assignment.time_slot}节')
            ws.cell(row=row, column=6, value=assignment.semester)
            ws.cell(row=row, column=7, value=assignment.academic_year)
            ws.cell(row=row, column=8, value=assignment.week_range)

        # 如果包含冲突信息
        if options.get('include_conflicts') and schedule_data.conflicts:
            # 添加冲突工作表
            ws_conflicts = wb.create_sheet("冲突信息")

            # 冲突表头
            conflict_headers = ['冲突类型', '描述', '严重性', '涉及课程']
            for col, header in enumerate(conflict_headers, 1):
                cell = ws_conflicts.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')

            # 填充冲突数据
            for row, conflict in enumerate(schedule_data.conflicts, 2):
                ws_conflicts.cell(row=row, column=1, value=conflict.conflict_type)
                ws_conflicts.cell(row=row, column=2, value=conflict.description)
                ws_conflicts.cell(row=row, column=3, value=conflict.severity)

                # 涉及的课程
                involved_courses = []
                for assignment in conflict.assignments:
                    course = course_dict.get(assignment.course_id, {})
                    involved_courses.append(course.get('name', f'Course {assignment.course_id}'))
                ws_conflicts.cell(row=row, column=4, value=', '.join(involved_courses))

        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    except ImportError:
        raise Exception("openpyxl库未安装，无法生成Excel文件")
    except Exception as e:
        raise Exception(f"生成Excel文件失败: {str(e)}")


def _generate_csv_export(schedule_data, options):
    """生成CSV导出文件"""
    try:
        import csv
        from io import StringIO

        output = StringIO()
        writer = csv.writer(output)

        # 写入表头
        headers = ['课程ID', '课程名称', '教师ID', '教师姓名', '教室ID', '教室名称',
                  '星期', '时间段', '学期', '学年', '周次']
        writer.writerow(headers)

        # 获取数据
        courses_data = _get_courses_data()
        teachers_data = _get_teachers_data()
        classrooms_data = _get_classrooms_data()

        course_dict = {c['id']: c for c in courses_data}
        teacher_dict = {t['id']: t for t in teachers_data}
        classroom_dict = {c['id']: c for c in classrooms_data}

        # 写入数据
        for assignment in schedule_data.assignments:
            course = course_dict.get(assignment.course_id, {})
            teacher = teacher_dict.get(assignment.teacher_id, {})
            classroom = classroom_dict.get(assignment.classroom_id, {})

            row = [
                assignment.course_id,
                course.get('name', f'Course {assignment.course_id}'),
                assignment.teacher_id,
                teacher.get('name', f'Teacher {assignment.teacher_id}'),
                assignment.classroom_id,
                classroom.get('name', f'Classroom {assignment.classroom_id}'),
                assignment.day_of_week,
                assignment.time_slot,
                assignment.semester,
                assignment.academic_year,
                assignment.week_range,
            ]
            writer.writerow(row)

        return output.getvalue()

    except Exception as e:
        raise Exception(f"生成CSV文件失败: {str(e)}")


def _generate_pdf_export(schedule_data, options):
    """生成PDF导出文件"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from io import BytesIO

        # 创建PDF文档
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)

        # 样式
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # 居中
        )

        # 内容列表
        story = []

        # 标题
        title = Paragraph("课程排课表", title_style)
        story.append(title)
        story.append(Spacer(1, 20))

        # 准备表格数据
        data = [['课程名称', '教师', '教室', '星期', '时间段', '学期', '学年']]

        courses_data = _get_courses_data()
        teachers_data = _get_teachers_data()
        classrooms_data = _get_classrooms_data()

        course_dict = {c['id']: c for c in courses_data}
        teacher_dict = {t['id']: t for t in teachers_data}
        classroom_dict = {c['id']: c for c in classrooms_data}

        day_names = ['', '周一', '周二', '周三', '周四', '周五', '周六', '周日']

        for assignment in schedule_data.assignments:
            course = course_dict.get(assignment.course_id, {})
            teacher = teacher_dict.get(assignment.teacher_id, {})
            classroom = classroom_dict.get(assignment.classroom_id, {})

            row = [
                course.get('name', f'Course {assignment.course_id}'),
                teacher.get('name', f'Teacher {assignment.teacher_id}'),
                classroom.get('name', f'Classroom {assignment.classroom_id}'),
                day_names[assignment.day_of_week] if assignment.day_of_week < len(day_names) else f'Day {assignment.day_of_week}',
                f'第{assignment.time_slot}节',
                assignment.semester,
                assignment.academic_year,
            ]
            data.append(row)

        # 创建表格
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        story.append(table)

        # 构建PDF
        doc.build(story)
        output.seek(0)

        return output.getvalue()

    except ImportError:
        raise Exception("reportlab库未安装，无法生成PDF文件")
    except Exception as e:
        raise Exception(f"生成PDF文件失败: {str(e)}")
