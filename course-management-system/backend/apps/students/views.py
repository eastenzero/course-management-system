from rest_framework import generics, status, permissions
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from django.contrib.auth import get_user_model
from django.http import HttpResponse
from django.db.models import Avg, Sum, Count, Q
from django.utils import timezone
from datetime import date, timedelta
from apps.courses.models import Course, Enrollment
from apps.users.permissions import IsStudent
from .models import StudentProfile, StudentCourseProgress
from .serializers import (
    StudentProfileSerializer, StudentDashboardSerializer,
    StudentEnrollmentSerializer, AvailableCourseSerializer,
    CourseScheduleSerializer, StudentCourseProgressSerializer
)
from .services import StudentService

User = get_user_model()


class StudentProfileView(generics.RetrieveUpdateAPIView):
    """学生档案视图"""
    
    serializer_class = StudentProfileSerializer
    permission_classes = [permissions.IsAuthenticated, IsStudent]
    
    def get_object(self):
        profile, created = StudentProfile.objects.get_or_create(
            user=self.request.user
        )
        return profile


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated, IsStudent])
def student_dashboard(request):
    """学生仪表板数据"""
    
    try:
        service = StudentService(request.user)
        dashboard_data = service.get_dashboard_data()
        
        serializer = StudentDashboardSerializer(dashboard_data)
        return Response(serializer.data)
    
    except Exception as e:
        return Response(
            {'error': f'获取仪表板数据失败: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated, IsStudent])
def available_courses(request):
    """获取可选课程列表"""
    
    try:
        # 获取查询参数
        semester = request.GET.get('semester')
        department = request.GET.get('department')
        course_type = request.GET.get('course_type')
        search = request.GET.get('search')
        
        # 构建查询条件
        queryset = Course.objects.filter(
            is_active=True,
            is_published=True
        )
        
        if semester:
            queryset = queryset.filter(semester=semester)
        if department:
            queryset = queryset.filter(department=department)
        if course_type:
            queryset = queryset.filter(course_type=course_type)
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(code__icontains=search) |
                Q(description__icontains=search)
            )
        
        # 排除已选课程
        enrolled_courses = Enrollment.objects.filter(
            student=request.user,
            status='enrolled'
        ).values_list('course_id', flat=True)
        
        queryset = queryset.exclude(id__in=enrolled_courses)
        
        # 序列化数据
        serializer = AvailableCourseSerializer(
            queryset,
            many=True,
            context={'request': request}
        )
        
        return Response(serializer.data)
    
    except Exception as e:
        return Response(
            {'error': f'获取可选课程失败: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated, IsStudent])
def enroll_course(request):
    """选课"""
    
    try:
        course_id = request.data.get('course_id')
        if not course_id:
            return Response(
                {'error': '课程ID不能为空'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        service = StudentService(request.user)
        result = service.enroll_course(course_id)
        
        if result['success']:
            return Response({
                'message': '选课成功',
                'enrollment': result['enrollment']
            })
        else:
            return Response(
                {'error': result['error']},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    except Exception as e:
        return Response(
            {'error': f'选课失败: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['DELETE'])
@permission_classes([permissions.IsAuthenticated, IsStudent])
def drop_course(request, course_id):
    """退课"""
    
    try:
        service = StudentService(request.user)
        result = service.drop_course(course_id)
        
        if result['success']:
            return Response({'message': '退课成功'})
        else:
            return Response(
                {'error': result['error']},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    except Exception as e:
        return Response(
            {'error': f'退课失败: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated, IsStudent])
def check_conflicts(request):
    """检查选课冲突"""
    
    try:
        course_ids = request.data.get('course_ids', [])
        if not course_ids:
            return Response(
                {'error': '课程ID列表不能为空'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        service = StudentService(request.user)
        conflicts = service.check_schedule_conflicts(course_ids)
        
        return Response({
            'has_conflicts': len(conflicts) > 0,
            'conflicts': conflicts
        })
    
    except Exception as e:
        return Response(
            {'error': f'冲突检测失败: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


class MyCoursesView(generics.ListAPIView):
    """我的课程列表"""
    
    serializer_class = StudentEnrollmentSerializer
    permission_classes = [permissions.IsAuthenticated, IsStudent]
    
    def get_queryset(self):
        return Enrollment.objects.filter(
            student=self.request.user,
            is_active=True
        ).select_related('course').order_by('-enrolled_at')


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated, IsStudent])
def course_schedule(request):
    """获取个人课程表"""
    
    try:
        semester = request.GET.get('semester')
        week = request.GET.get('week')
        
        service = StudentService(request.user)
        schedule_data = service.get_course_schedule(semester, week)
        
        serializer = CourseScheduleSerializer(schedule_data, many=True)
        return Response(serializer.data)
    
    except Exception as e:
        return Response(
            {'error': f'获取课程表失败: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated, IsStudent])
def grades_list(request):
    """成绩列表"""
    
    try:
        semester = request.GET.get('semester')
        academic_year = request.GET.get('academic_year')
        
        queryset = Enrollment.objects.filter(
            student=request.user,
            is_active=True
        ).select_related('course')
        
        if semester:
            queryset = queryset.filter(course__semester=semester)
        if academic_year:
            queryset = queryset.filter(course__academic_year=academic_year)
        
        serializer = StudentEnrollmentSerializer(queryset, many=True)
        return Response(serializer.data)
    
    except Exception as e:
        return Response(
            {'error': f'获取成绩失败: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated, IsStudent])
def gpa_statistics(request):
    """获取GPA统计"""
    
    try:
        service = StudentService(request.user)
        gpa_data = service.get_gpa_statistics()
        
        return Response(gpa_data)
    
    except Exception as e:
        return Response(
            {'error': f'获取GPA统计失败: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated, IsStudent])
def export_schedule(request):
    """导出学生课程表"""
    
    try:
        semester = request.GET.get('semester')
        format_type = request.GET.get('format', 'excel')
        
        # 获取学生课程表数据
        service = StudentService(request.user)
        schedule_data = service.get_course_schedule(semester)
        
        if not schedule_data:
            return Response(
                {'error': '暂无课程表数据'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # 根据格式生成导出文件
        if format_type == 'excel':
            file_content = _generate_schedule_excel(schedule_data, request.user, semester)
            response = HttpResponse(
                file_content,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f'课程表_{request.user.username}_{semester or "全部"}.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        elif format_type == 'csv':
            file_content = _generate_schedule_csv(schedule_data, request.user, semester)
            response = HttpResponse(file_content, content_type='text/csv; charset=utf-8')
            filename = f'课程表_{request.user.username}_{semester or "全部"}.csv'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        elif format_type == 'pdf':
            file_content = _generate_schedule_pdf(schedule_data, request.user, semester)
            response = HttpResponse(file_content, content_type='application/pdf')
            filename = f'课程表_{request.user.username}_{semester or "全部"}.pdf'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        else:
            return Response(
                {'error': f'不支持的导出格式: {format_type}'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    except Exception as e:
        return Response(
            {'error': f'导出课程表失败: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated, IsStudent])
def export_grades(request):
    """导出学生成绩单"""
    
    try:
        semester = request.GET.get('semester')
        academic_year = request.GET.get('academic_year')
        format_type = request.GET.get('format', 'excel')
        
        # 获取成绩数据
        queryset = Enrollment.objects.filter(
            student=request.user,
            is_active=True,
            score__isnull=False
        ).select_related('course')
        
        if semester:
            queryset = queryset.filter(course__semester=semester)
        if academic_year:
            queryset = queryset.filter(course__academic_year=academic_year)
        
        grades_data = StudentEnrollmentSerializer(queryset, many=True).data
        
        if not grades_data:
            return Response(
                {'error': '暂无成绩数据'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # 根据格式生成导出文件
        if format_type == 'excel':
            file_content = _generate_grades_excel(grades_data, request.user, semester)
            response = HttpResponse(
                file_content,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f'成绩单_{request.user.username}_{semester or "全部"}.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        elif format_type == 'csv':
            file_content = _generate_grades_csv(grades_data, request.user, semester)
            response = HttpResponse(file_content, content_type='text/csv; charset=utf-8')
            filename = f'成绩单_{request.user.username}_{semester or "全部"}.csv'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        else:
            return Response(
                {'error': f'不支持的导出格式: {format_type}'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    except Exception as e:
        return Response(
            {'error': f'导出成绩单失败: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


def _generate_schedule_excel(schedule_data, user, semester):
    """生成课程表Excel文件"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("需要安装 openpyxl 库来支持Excel导出")
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"个人课程表-{user.username}"
    
    # 设置样式
    title_font = Font(size=16, bold=True)
    header_font = Font(size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 标题
    ws['A1'] = f"个人课程表 - {user.get_full_name() or user.username}"
    ws['A1'].font = title_font
    ws.merge_cells('A1:H1')
    
    ws['A2'] = f"学期：{semester or '全部'}" 
    ws.merge_cells('A2:H2')
    
    # 表头
    headers = [
        '课程代码', '课程名称', '授课教师', '教室',
        '星期', '时间段', '上课时间', '周次'
    ]
    
    row = 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # 数据行
    for item in schedule_data:
        row += 1
        data_row = [
            item['course_code'],
            item['course_name'], 
            item['teacher_name'],
            item['classroom'],
            item['day_of_week_display'],
            item['time_slot'],
            f"{item['start_time']}-{item['end_time']}",
            item['week_range']
        ]
        
        for col, value in enumerate(data_row, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 调整列宽
    column_widths = [12, 20, 15, 12, 8, 15, 15, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # 保存到内存
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _generate_schedule_csv(schedule_data, user, semester):
    """生成课程表CSV文件"""
    import csv
    from io import StringIO
    
    output = StringIO()
    writer = csv.writer(output)
    
    # 写入标题
    writer.writerow([f"个人课程表 - {user.get_full_name() or user.username}"])
    writer.writerow([f"学期：{semester or '全部'}"])
    writer.writerow([])  # 空行
    
    # 写入表头
    writer.writerow([
        '课程代码', '课程名称', '授课教师', '教室',
        '星期', '时间段', '上课时间', '周次'
    ])
    
    # 写入数据
    for item in schedule_data:
        writer.writerow([
            item['course_code'],
            item['course_name'],
            item['teacher_name'],
            item['classroom'],
            item['day_of_week_display'],
            item['time_slot'],
            f"{item['start_time']}-{item['end_time']}",
            item['week_range']
        ])
    
    return output.getvalue().encode('utf-8-sig')


def _generate_schedule_pdf(schedule_data, user, semester):
    """生成课程表PDF文件"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from io import BytesIO
    except ImportError:
        raise ImportError("需要安装 reportlab 库来支持PDF导出")
    
    buffer = BytesIO()
    
    # 创建PDF文档
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # 样式
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # 居中
    )
    
    # 标题
    title = Paragraph(f"个人课程表 - {user.get_full_name() or user.username}", title_style)
    elements.append(title)
    
    subtitle = Paragraph(f"学期：{semester or '全部'}", styles['Normal'])
    elements.append(subtitle)
    elements.append(Spacer(1, 20))
    
    # 表格数据
    table_data = [
        ['课程代码', '课程名称', '教师', '教室', '星期', '时间', '周次']
    ]
    
    for item in schedule_data:
        table_data.append([
            item['course_code'],
            item['course_name'][:10] + '...' if len(item['course_name']) > 10 else item['course_name'],
            item['teacher_name'],
            item['classroom'],
            item['day_of_week_display'],
            f"{item['start_time']}-{item['end_time']}",
            item['week_range']
        ])
    
    # 创建表格
    table = Table(table_data, colWidths=[1*inch, 1.5*inch, 1*inch, 1*inch, 0.8*inch, 1*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    
    # 生成PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def _generate_grades_excel(grades_data, user, semester):
    """生成成绩单Excel文件"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("需要安装 openpyxl 库来支持Excel导出")
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"成绩单-{user.username}"
    
    # 设置样式
    title_font = Font(size=16, bold=True)
    header_font = Font(size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 标题
    ws['A1'] = f"成绩单 - {user.get_full_name() or user.username}"
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')
    
    ws['A2'] = f"学期：{semester or '全部'}"
    ws.merge_cells('A2:F2')
    
    # 表头
    headers = ['课程代码', '课程名称', '学分', '成绩', '等级', '绩点']
    
    row = 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # 数据行
    for item in grades_data:
        row += 1
        course_info = item.get('course_info', {})
        data_row = [
            course_info.get('code', ''),
            course_info.get('name', ''),
            course_info.get('credits', ''),
            item.get('score', ''),
            item.get('grade', ''),
            # 简单的绩点计算
            '4.0' if float(item.get('score', 0)) >= 90 else 
            '3.0' if float(item.get('score', 0)) >= 80 else
            '2.0' if float(item.get('score', 0)) >= 70 else
            '1.0' if float(item.get('score', 0)) >= 60 else '0.0'
        ]
        
        for col, value in enumerate(data_row, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 调整列宽
    column_widths = [15, 25, 8, 8, 8, 8]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # 保存到内存
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _generate_grades_csv(grades_data, user, semester):
    """生成成绩单CSV文件"""
    import csv
    from io import StringIO
    
    output = StringIO()
    writer = csv.writer(output)
    
    # 写入标题
    writer.writerow([f"成绩单 - {user.get_full_name() or user.username}"])
    writer.writerow([f"学期：{semester or '全部'}"])
    writer.writerow([])  # 空行
    
    # 写入表头
    writer.writerow(['课程代码', '课程名称', '学分', '成绩', '等级', '绩点'])
    
    # 写入数据
    for item in grades_data:
        course_info = item.get('course_info', {})
        score = float(item.get('score', 0))
        gpa = '4.0' if score >= 90 else '3.0' if score >= 80 else '2.0' if score >= 70 else '1.0' if score >= 60 else '0.0'
        
        writer.writerow([
            course_info.get('code', ''),
            course_info.get('name', ''),
            course_info.get('credits', ''),
            item.get('score', ''),
            item.get('grade', ''),
            gpa
        ])
    
    return output.getvalue().encode('utf-8-sig')
