"""
课程表相关服务类
"""

from django.db.models import Q
from django.core.exceptions import ValidationError
from django.utils import timezone
from typing import List, Dict, Any, Optional, Tuple
from .models import Schedule, TimeSlot
from apps.courses.models import Course
from apps.classrooms.models import Classroom
from django.contrib.auth import get_user_model

User = get_user_model()


class ScheduleConflictDetector:
    """课程表冲突检测器"""
    
    @staticmethod
    def detect_all_conflicts(semester: str) -> Dict[str, List[Dict]]:
        """检测指定学期的所有冲突
        
        Args:
            semester: 学期
            
        Returns:
            dict: 冲突报告
        """
        conflicts = {
            'classroom_conflicts': [],
            'teacher_conflicts': [],
            'time_conflicts': [],
            'capacity_conflicts': []
        }
        
        schedules = Schedule.objects.filter(
            semester=semester,
            status='active'
        ).select_related('course', 'teacher', 'classroom', 'time_slot')
        
        # 检测教室冲突
        classroom_conflicts = ScheduleConflictDetector._detect_classroom_conflicts(schedules)
        conflicts['classroom_conflicts'] = classroom_conflicts
        
        # 检测教师冲突
        teacher_conflicts = ScheduleConflictDetector._detect_teacher_conflicts(schedules)
        conflicts['teacher_conflicts'] = teacher_conflicts
        
        # 检测时间冲突
        time_conflicts = ScheduleConflictDetector._detect_time_conflicts(schedules)
        conflicts['time_conflicts'] = time_conflicts
        
        # 检测容量冲突
        capacity_conflicts = ScheduleConflictDetector._detect_capacity_conflicts(schedules)
        conflicts['capacity_conflicts'] = capacity_conflicts
        
        return conflicts
    
    @staticmethod
    def _detect_classroom_conflicts(schedules) -> List[Dict]:
        """检测教室冲突"""
        conflicts = []
        classroom_schedule_map = {}
        
        for schedule in schedules:
            key = (
                schedule.classroom.id,
                schedule.day_of_week,
                schedule.time_slot.id
            )
            
            if key not in classroom_schedule_map:
                classroom_schedule_map[key] = []
            classroom_schedule_map[key].append(schedule)
        
        # 找出有冲突的教室
        for key, schedule_list in classroom_schedule_map.items():
            if len(schedule_list) > 1:
                conflicts.append({
                    'type': 'classroom_conflict',
                    'classroom': str(schedule_list[0].classroom),
                    'day_of_week': schedule_list[0].get_day_of_week_display(),
                    'time_slot': schedule_list[0].time_slot.name,
                    'conflicting_schedules': [
                        {
                            'id': s.id,
                            'course': s.course.name,
                            'teacher': s.teacher.get_full_name() or s.teacher.username,
                        }
                        for s in schedule_list
                    ]
                })
        
        return conflicts
    
    @staticmethod
    def _detect_teacher_conflicts(schedules) -> List[Dict]:
        """检测教师冲突"""
        conflicts = []
        teacher_schedule_map = {}
        
        for schedule in schedules:
            key = (
                schedule.teacher.id,
                schedule.day_of_week,
                schedule.time_slot.id
            )
            
            if key not in teacher_schedule_map:
                teacher_schedule_map[key] = []
            teacher_schedule_map[key].append(schedule)
        
        # 找出有冲突的教师
        for key, schedule_list in teacher_schedule_map.items():
            if len(schedule_list) > 1:
                conflicts.append({
                    'type': 'teacher_conflict',
                    'teacher': schedule_list[0].teacher.get_full_name() or schedule_list[0].teacher.username,
                    'day_of_week': schedule_list[0].get_day_of_week_display(),
                    'time_slot': schedule_list[0].time_slot.name,
                    'conflicting_schedules': [
                        {
                            'id': s.id,
                            'course': s.course.name,
                            'classroom': str(s.classroom),
                        }
                        for s in schedule_list
                    ]
                })
        
        return conflicts
    
    @staticmethod
    def _detect_time_conflicts(schedules) -> List[Dict]:
        """检测时间冲突（重叠的时间段）"""
        conflicts = []
        time_slots = TimeSlot.objects.filter(is_active=True).order_by('start_time')
        
        # 检查时间段是否有重叠
        for i, slot1 in enumerate(time_slots):
            for slot2 in time_slots[i+1:]:
                if ScheduleConflictDetector._time_slots_overlap(slot1, slot2):
                    # 找出使用这两个时间段的课程
                    schedules1 = schedules.filter(time_slot=slot1)
                    schedules2 = schedules.filter(time_slot=slot2)
                    
                    if schedules1.exists() and schedules2.exists():
                        conflicts.append({
                            'type': 'time_overlap',
                            'time_slot_1': slot1.name,
                            'time_slot_2': slot2.name,
                            'overlap_period': f"{max(slot1.start_time, slot2.start_time)}-{min(slot1.end_time, slot2.end_time)}",
                            'affected_schedules': {
                                'slot1_schedules': [s.course.name for s in schedules1],
                                'slot2_schedules': [s.course.name for s in schedules2],
                            }
                        })
        
        return conflicts
    
    @staticmethod
    def _detect_capacity_conflicts(schedules) -> List[Dict]:
        """检测容量冲突"""
        conflicts = []
        
        for schedule in schedules:
            if schedule.classroom.capacity < schedule.course.max_students:
                conflicts.append({
                    'type': 'capacity_conflict',
                    'schedule_id': schedule.id,
                    'course': schedule.course.name,
                    'classroom': str(schedule.classroom),
                    'classroom_capacity': schedule.classroom.capacity,
                    'course_max_students': schedule.course.max_students,
                    'shortage': schedule.course.max_students - schedule.classroom.capacity,
                })
        
        return conflicts
    
    @staticmethod
    def _time_slots_overlap(slot1: TimeSlot, slot2: TimeSlot) -> bool:
        """检查两个时间段是否重叠"""
        return (slot1.start_time < slot2.end_time and 
                slot2.start_time < slot1.end_time)


class ScheduleOptimizer:
    """课程表优化器"""
    
    @staticmethod
    def suggest_alternative_schedules(
        course: Course,
        teacher: User,
        semester: str,
        preferred_days: List[int] = None,
        preferred_times: List[int] = None
    ) -> List[Dict]:
        """为课程建议可选的时间安排
        
        Args:
            course: 课程
            teacher: 教师
            semester: 学期
            preferred_days: 偏好的星期（1-7）
            preferred_times: 偏好的时间段ID
            
        Returns:
            list: 可选的时间安排
        """
        suggestions = []
        
        # 获取所有可用的时间段
        time_slots = TimeSlot.objects.filter(is_active=True).order_by('order')
        
        # 获取合适容量的教室
        suitable_classrooms = Classroom.objects.filter(
            capacity__gte=course.max_students,
            is_active=True
        ).order_by('capacity')
        
        days = preferred_days or [1, 2, 3, 4, 5]  # 默认工作日
        time_slot_ids = preferred_times or [ts.id for ts in time_slots]
        
        for day in days:
            for time_slot in time_slots.filter(id__in=time_slot_ids):
                for classroom in suitable_classrooms:
                    # 检查是否有冲突
                    classroom_conflicts = Schedule.check_classroom_conflicts(
                        classroom, day, time_slot, semester
                    )
                    teacher_conflicts = Schedule.check_teacher_conflicts(
                        teacher, day, time_slot, semester
                    )
                    
                    if not classroom_conflicts.exists() and not teacher_conflicts.exists():
                        suggestions.append({
                            'day_of_week': day,
                            'day_name': dict(Schedule.DAY_CHOICES)[day],
                            'time_slot': {
                                'id': time_slot.id,
                                'name': time_slot.name,
                                'start_time': time_slot.start_time.strftime('%H:%M'),
                                'end_time': time_slot.end_time.strftime('%H:%M'),
                            },
                            'classroom': {
                                'id': classroom.id,
                                'name': str(classroom),
                                'capacity': classroom.capacity,
                                'building': classroom.building.name,
                            },
                            'score': ScheduleOptimizer._calculate_schedule_score(
                                day, time_slot, classroom, course, teacher
                            )
                        })
        
        # 按评分排序
        suggestions.sort(key=lambda x: x['score'], reverse=True)
        
        return suggestions[:10]  # 返回前10个建议
    
    @staticmethod
    def _calculate_schedule_score(
        day: int,
        time_slot: TimeSlot,
        classroom: Classroom,
        course: Course,
        teacher: User
    ) -> float:
        """计算时间安排的评分
        
        评分标准：
        - 工作日优于周末
        - 上午和下午优于晚上
        - 教室容量适中（不要太大也不要太小）
        - 教室类型匹配课程类型
        """
        score = 0.0
        
        # 星期评分
        if day <= 5:  # 工作日
            score += 10
        else:  # 周末
            score += 5
        
        # 时间段评分
        hour = time_slot.start_time.hour
        if 8 <= hour <= 11:  # 上午
            score += 10
        elif 14 <= hour <= 17:  # 下午
            score += 8
        elif 19 <= hour <= 21:  # 晚上
            score += 6
        else:  # 其他时间
            score += 3
        
        # 教室容量评分
        capacity_ratio = classroom.capacity / course.max_students
        if 1.0 <= capacity_ratio <= 1.5:  # 容量刚好合适
            score += 10
        elif 1.5 < capacity_ratio <= 2.0:  # 容量稍大
            score += 8
        elif capacity_ratio > 2.0:  # 容量过大
            score += 5
        
        # 教室类型匹配评分
        if hasattr(classroom, 'room_type') and hasattr(course, 'course_type'):
            if (classroom.room_type == 'lab' and 'lab' in course.name.lower()) or \
               (classroom.room_type == 'lecture' and course.course_type in ['required', 'elective']):
                score += 5
        
        return score


class ScheduleImportExportService:
    """课程表导入导出服务"""
    
    @staticmethod
    def export_schedule_to_dict(semester: str) -> Dict[str, Any]:
        """导出课程表为字典格式"""
        schedules = Schedule.objects.filter(
            semester=semester,
            status='active'
        ).select_related(
            'course', 'teacher', 'classroom', 'time_slot'
        ).order_by('day_of_week', 'time_slot__order')
        
        export_data = {
            'semester': semester,
            'export_time': str(timezone.now()),
            'schedules': []
        }
        
        for schedule in schedules:
            export_data['schedules'].append({
                'course_code': schedule.course.code,
                'course_name': schedule.course.name,
                'teacher_id': schedule.teacher.employee_id if hasattr(schedule.teacher, 'employee_id') else schedule.teacher.username,
                'teacher_name': schedule.teacher.get_full_name() or schedule.teacher.username,
                'classroom_code': getattr(schedule.classroom, 'code', str(schedule.classroom)),
                'classroom_name': str(schedule.classroom),
                'day_of_week': schedule.day_of_week,
                'time_slot_name': schedule.time_slot.name,
                'start_time': schedule.time_slot.start_time.strftime('%H:%M'),
                'end_time': schedule.time_slot.end_time.strftime('%H:%M'),
                'week_range': schedule.week_range,
                'notes': schedule.notes,
            })
        
        return export_data
    
    @staticmethod
    def validate_import_data(data: Dict[str, Any]) -> Tuple[bool, List[str]]:
        """验证导入数据"""
        errors = []
        
        if 'semester' not in data:
            errors.append('缺少学期信息')
        
        if 'schedules' not in data or not isinstance(data['schedules'], list):
            errors.append('缺少课程安排数据或格式错误')
            return False, errors
        
        required_fields = [
            'course_code', 'teacher_id', 'classroom_code',
            'day_of_week', 'time_slot_name', 'week_range'
        ]
        
        for i, schedule_data in enumerate(data['schedules']):
            for field in required_fields:
                if field not in schedule_data:
                    errors.append(f'第{i+1}行缺少必需字段: {field}')
            
            # 验证星期
            if 'day_of_week' in schedule_data:
                try:
                    day = int(schedule_data['day_of_week'])
                    if not 1 <= day <= 7:
                        errors.append(f'第{i+1}行星期值无效: {day}')
                except (ValueError, TypeError):
                    errors.append(f'第{i+1}行星期值格式错误')
        
        return len(errors) == 0, errors

    @staticmethod
    def import_schedules_from_data(data: Dict[str, Any]) -> Tuple[int, List[str]]:
        """从导入数据创建课程安排

        Args:
            data: 导入数据

        Returns:
            tuple: (成功创建数量, 错误列表)
        """
        semester = data.get('semester')
        schedules_data = data.get('schedules', [])

        created_count = 0
        errors = []

        for i, schedule_data in enumerate(schedules_data):
            try:
                # 查找课程
                course_code = schedule_data.get('course_code')
                course = Course.objects.filter(code=course_code).first()
                if not course:
                    errors.append(f'第{i+1}行: 找不到课程代码 {course_code}')
                    continue

                # 查找教师
                teacher_id = schedule_data.get('teacher_id')
                teacher = User.objects.filter(
                    Q(employee_id=teacher_id) | Q(username=teacher_id),
                    user_type='teacher'
                ).first()
                if not teacher:
                    errors.append(f'第{i+1}行: 找不到教师 {teacher_id}')
                    continue

                # 查找教室
                classroom_code = schedule_data.get('classroom_code')
                classroom = Classroom.objects.filter(
                    Q(code=classroom_code) | Q(name=classroom_code)
                ).first()
                if not classroom:
                    errors.append(f'第{i+1}行: 找不到教室 {classroom_code}')
                    continue

                # 查找时间段
                time_slot_name = schedule_data.get('time_slot_name')
                start_time = schedule_data.get('start_time')
                end_time = schedule_data.get('end_time')

                time_slot = TimeSlot.objects.filter(
                    Q(name=time_slot_name) |
                    Q(start_time=start_time, end_time=end_time)
                ).first()
                if not time_slot:
                    errors.append(f'第{i+1}行: 找不到时间段 {time_slot_name}')
                    continue

                # 创建课程安排
                schedule = Schedule(
                    course=course,
                    teacher=teacher,
                    classroom=classroom,
                    time_slot=time_slot,
                    day_of_week=schedule_data.get('day_of_week'),
                    week_range=schedule_data.get('week_range', '1-16周'),
                    semester=semester,
                    academic_year=semester.split('-')[0] + '-' + semester.split('-')[1] if semester else '',
                    notes=schedule_data.get('notes', '')
                )

                # 验证并保存
                schedule.full_clean()
                schedule.save()
                created_count += 1

            except ValidationError as e:
                errors.append(f'第{i+1}行: 验证错误 - {str(e)}')
            except Exception as e:
                errors.append(f'第{i+1}行: 创建失败 - {str(e)}')

        return created_count, errors

    @staticmethod
    def export_schedule_to_excel(semester: str, options: Dict[str, Any] = None) -> bytes:
        """导出课程表为Excel格式

        Args:
            semester: 学期
            options: 导出选项

        Returns:
            bytes: Excel文件内容
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("需要安装 openpyxl 库来支持Excel导出")

        options = options or {}
        include_weekend = options.get('include_weekend', False)
        group_by = options.get('group_by', 'teacher')

        wb = openpyxl.Workbook()

        schedules = Schedule.objects.filter(
            semester=semester,
            status='active'
        ).select_related(
            'course', 'teacher', 'classroom', 'time_slot'
        ).order_by('day_of_week', 'time_slot__order')

        if group_by in ['teacher', 'classroom']:
            if 'Sheet' in wb.sheetnames and len(wb.sheetnames) == 1:
                ws_to_remove = wb.active
                wb.remove(ws_to_remove)

            groups = {}
            for s in schedules:
                key = (s.teacher.get_full_name() or s.teacher.username) if group_by == 'teacher' else str(s.classroom)
                groups.setdefault(key, []).append(s)

            for name, sched_list in groups.items():
                title = f"{('教师' if group_by=='teacher' else '教室')}-{name}"[:31]
                ws = wb.create_sheet(title)
                table = ScheduleImportExportService._create_schedule_table_data(sched_list, include_weekend)
                for r_idx, row_vals in enumerate(table, start=1):
                    for c_idx, val in enumerate(row_vals, start=1):
                        ws.cell(row=r_idx, column=c_idx, value=val)
                for col in range(1, len(table[0]) + 1):
                    ws.column_dimensions[get_column_letter(col)].width = 18
        else:
            ws = wb.active
            ws.title = f"课程表-{semester}"
            headers = [
                '课程代码', '课程名称', '教师姓名', '教室',
                '星期', '时间段', '开始时间', '结束时间',
                '周次范围', '备注'
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            row = 2
            for schedule in schedules:
                ws.cell(row=row, column=1, value=schedule.course.code)
                ws.cell(row=row, column=2, value=schedule.course.name)
                ws.cell(row=row, column=3, value=schedule.teacher.get_full_name() or schedule.teacher.username)
                ws.cell(row=row, column=4, value=str(schedule.classroom))
                ws.cell(row=row, column=5, value=schedule.get_day_of_week_display())
                ws.cell(row=row, column=6, value=schedule.time_slot.name)
                ws.cell(row=row, column=7, value=schedule.time_slot.start_time.strftime('%H:%M'))
                ws.cell(row=row, column=8, value=schedule.time_slot.end_time.strftime('%H:%M'))
                ws.cell(row=row, column=9, value=schedule.week_range)
                ws.cell(row=row, column=10, value=schedule.notes)
                row += 1
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15

        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    @staticmethod
    def export_schedule_to_csv(semester: str, options: Dict[str, Any] = None) -> str:
        """导出课程表为CSV格式

        Args:
            semester: 学期
            options: 导出选项

        Returns:
            str: CSV文件内容
        """
        import csv
        from io import StringIO

        options = options or {}

        # 获取课程安排数据
        schedules = Schedule.objects.filter(
            semester=semester,
            status='active'
        ).select_related(
            'course', 'teacher', 'classroom', 'time_slot'
        ).order_by('day_of_week', 'time_slot__order')

        # 创建CSV内容
        output = StringIO()
        writer = csv.writer(output)

        # 写入标题行
        headers = [
            '课程代码', '课程名称', '教师姓名', '教室',
            '星期', '时间段', '开始时间', '结束时间',
            '周次范围', '备注'
        ]
        writer.writerow(headers)

        # 写入数据
        for schedule in schedules:
            writer.writerow([
                schedule.course.code,
                schedule.course.name,
                schedule.teacher.get_full_name() or schedule.teacher.username,
                str(schedule.classroom),
                schedule.get_day_of_week_display(),
                schedule.time_slot.name,
                schedule.time_slot.start_time.strftime('%H:%M'),
                schedule.time_slot.end_time.strftime('%H:%M'),
                schedule.week_range,
                schedule.notes
            ])

        return output.getvalue()

    @staticmethod
    def export_schedule_to_pdf(semester: str, options: Dict[str, Any] = None) -> bytes:
        """导出课程表为PDF格式

        Args:
            semester: 学期
            options: 导出选项

        Returns:
            bytes: PDF文件内容
        """
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from io import BytesIO
        except ImportError:
            raise ImportError("需要安装 reportlab 库来支持PDF导出")

        options = options or {}
        include_weekend = options.get('include_weekend', False)
        group_by = options.get('group_by', 'teacher')

        # 创建PDF文档
        output = BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )

        # 样式设置
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=20,
            alignment=1,  # 居中
            textColor=colors.darkblue
        )

        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=15,
            alignment=1,  # 居中
            textColor=colors.grey
        )

        # 获取课程安排数据
        schedules = Schedule.objects.filter(
            semester=semester,
            status='active'
        ).select_related(
            'course', 'teacher', 'classroom', 'time_slot'
        ).order_by('day_of_week', 'time_slot__order')

        # 内容列表
        story = []

        # 标题
        title = Paragraph(f"课程表 - {semester}", title_style)
        story.append(title)

        subtitle = Paragraph(f"生成时间: {timezone.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style)
        story.append(subtitle)
        story.append(Spacer(1, 20))

        if group_by == 'teacher':
            # 按教师分组
            teacher_schedules = {}
            for schedule in schedules:
                teacher_name = schedule.teacher.get_full_name() or schedule.teacher.username
                if teacher_name not in teacher_schedules:
                    teacher_schedules[teacher_name] = []
                teacher_schedules[teacher_name].append(schedule)

            for teacher_name, teacher_schedule_list in teacher_schedules.items():
                # 教师名称
                teacher_title = Paragraph(f"教师: {teacher_name}", styles['Heading2'])
                story.append(teacher_title)
                story.append(Spacer(1, 10))

                # 创建课程表格
                table_data = ScheduleImportExportService._create_schedule_table_data(
                    teacher_schedule_list, include_weekend
                )

                if table_data:
                    table = Table(table_data, colWidths=[1.2*inch, 1*inch, 1*inch, 1*inch, 1*inch, 1*inch, 1*inch])
                    table.setStyle(ScheduleImportExportService._get_table_style())
                    story.append(table)
                    story.append(Spacer(1, 20))

        elif group_by == 'classroom':
            # 按教室分组
            classroom_schedules = {}
            for schedule in schedules:
                classroom_name = str(schedule.classroom)
                if classroom_name not in classroom_schedules:
                    classroom_schedules[classroom_name] = []
                classroom_schedules[classroom_name].append(schedule)

            for classroom_name, classroom_schedule_list in classroom_schedules.items():
                # 教室名称
                classroom_title = Paragraph(f"教室: {classroom_name}", styles['Heading2'])
                story.append(classroom_title)
                story.append(Spacer(1, 10))

                # 创建课程表格
                table_data = ScheduleImportExportService._create_schedule_table_data(
                    classroom_schedule_list, include_weekend
                )

                if table_data:
                    table = Table(table_data, colWidths=[1.2*inch, 1*inch, 1*inch, 1*inch, 1*inch, 1*inch, 1*inch])
                    table.setStyle(ScheduleImportExportService._get_table_style())
                    story.append(table)
                    story.append(Spacer(1, 20))

        else:
            # 整体课程表
            # 准备表格数据
            headers = ['课程名称', '教师', '教室', '星期', '时间段', '开始时间', '结束时间']
            data = [headers]

            for schedule in schedules:
                row = [
                    schedule.course.name,
                    schedule.teacher.get_full_name() or schedule.teacher.username,
                    str(schedule.classroom),
                    schedule.get_day_of_week_display(),
                    schedule.time_slot.name,
                    schedule.time_slot.start_time.strftime('%H:%M'),
                    schedule.time_slot.end_time.strftime('%H:%M')
                ]
                data.append(row)

            # 创建表格
            if len(data) > 1:
                table = Table(data, colWidths=[1.5*inch, 1*inch, 1*inch, 0.8*inch, 1*inch, 0.8*inch, 0.8*inch])
                table.setStyle(ScheduleImportExportService._get_table_style())
                story.append(table)

        # 构建PDF
        doc.build(story)
        output.seek(0)

        return output.getvalue()

    @staticmethod
    def _create_schedule_table_data(schedules: List[Schedule], include_weekend: bool = False) -> List[List[str]]:
        active_slots = TimeSlot.objects.filter(is_active=True).order_by('order')
        days = list(range(1, 8 if include_weekend else 6))
        day_names = ['', '周一', '周二', '周三', '周四', '周五', '周六', '周日']
        headers = ['时间段'] + [day_names[day] for day in days]
        data = [headers]

        schedule_map: Dict[Tuple[int, int], List[Schedule]] = {}
        for schedule in schedules:
            key = (schedule.day_of_week, schedule.time_slot.id)
            schedule_map.setdefault(key, []).append(schedule)

        for ts in active_slots:
            row = [ts.name]
            for day in days:
                key = (day, ts.id)
                if key in schedule_map:
                    texts = []
                    for s in schedule_map[key]:
                        texts.append(f"{s.course.name}\n{str(s.classroom)}")
                    row.append('\n'.join(texts))
                else:
                    row.append('')
            data.append(row)

        return data

    @staticmethod
    def _get_table_style():
        """获取表格样式"""
        from reportlab.platypus import TableStyle
        from reportlab.lib import colors

        return TableStyle([
            # 表头样式
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

            # 数据行样式
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

            # 交替行颜色
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.lightgrey, colors.white]),
        ])
